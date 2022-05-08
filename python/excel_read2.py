import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook
import pymysql
import sys
# Open the workbook and select a worksheet
path = sys.argv[1]
print(path)
wb = load_workbook(path, data_only=True)
sheet = wb['ChatbotData4']
# List to hold dictionaries

#conn = pymysql.connect(host='61.42.251.243', user='playcom', password='tmvlfldzm!23',
#                       db='spdb', port=39306, charset='utf8')
conn = pymysql.connect(host='192.168.180.236', user='playcom', password='tmvlfldzm!23',
                       db='spdb', port=39306, charset='utf8')


curs = conn.cursor()
sql = """insert into tb_work_test2(quest,answer,label) values (%s, %s, %s)"""

# Iterate through each row in worksheet and fetch values into dict
for row in islice(sheet.values, 1, sheet.max_row):
    print(row[0])
    curs.execute(sql, (row[0],row[1],row[2]))
conn.commit()
conn.close()
# Serialize the list of dicts to JSON
