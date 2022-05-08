from openpyxl import load_workbook

 
#data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("C:/project/temp/업로드테스트_202104160452.xlsx", data_only=True)
#시트 이름으로 불러오기
load_ws = load_wb['Sheet1']
 
#셀 주소로 값 출력
#print(load_ws['A1'].value)
 
#셀 좌표로 값 출력
#print(load_ws.cell(1,2).value)

print('\n-----모든 행과 열 출력-----')
all_values = []
for row in load_ws.rows:
    row_value = []
    for cell in row:
        row_value.append(cell.value)
    all_values.append(row_value)
print(all_values)
