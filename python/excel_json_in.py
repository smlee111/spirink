import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook
import pymysql
import sys
# Open the workbook and select a worksheet
path = sys.argv[1]
bd_idx = sys.argv[2]
db_name = sys.argv[3]
print(path)
wb = load_workbook(path, data_only=True)
sheet = wb.active
# List to hold dictionaries

#conn = pymysql.connect(host='61.42.251.243', user='playcom', password='tmvlfldzm!23',
#                       db='spdb', port=39306, charset='utf8')
conn = pymysql.connect(host='192.168.180.236', user='playcom', password='tmvlfldzm!23',
                       db=db_name, port=39306, charset='utf8')
#데이터 입력
table_name = 'tb_json_'+bd_idx
curs = conn.cursor()
sql = "insert into {0}(uni_num,bd_idx,quest) values (%s, %s, %s)".format(table_name)

# Iterate through each row in worksheet and fetch values into dict
for row in islice(sheet.values, 1, sheet.max_row):
    print(row[0])
    curs.execute(sql, (row[0],bd_idx,row[1]))
#마스터 업데이트
sql_update = "update tb_work_bd set bd_cnt=%s where bd_idx=%s"
curs.execute(sql_update, (sheet.max_row-1,bd_idx))

conn.commit()
conn.close()
# Serialize the list of dicts to JSON
